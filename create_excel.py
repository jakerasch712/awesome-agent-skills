#!/usr/bin/env python3
"""Parse README.md and generate a categorized skills.xlsx file."""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

README = "README.md"
OUTPUT = "skills.xlsx"

# ── Regex patterns ────────────────────────────────────────────────────────────
SKILL_RE = re.compile(r'^\s*-\s+\*\*\[([^\]]+)\]\(([^)]+)\)\*\*\s*(?:-\s*)?(.*)', re.DOTALL)
DETAILS_H3_RE = re.compile(r'<summary><h3[^>]*>([^<]+)</h3></summary>')
PLAIN_H3_RE = re.compile(r'^###\s+(.+)')

# Microsoft language sub-sections that should be prefixed
MICROSOFT_SUBSECTIONS = {".NET Skills", "Java Skills", "Python Skills",
                         "Rust Skills", "TypeScript Skills", "General Skills"}

# ── Styling constants ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")
TOTAL_FONT  = Font(bold=True, size=11)
TOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
URL_FONT    = Font(color="0563C1", underline="single")
THIN_BORDER = Border(
    bottom=Side(border_style="thin", color="B8CCE4")
)

SHEET_COLS = ["#", "Category", "Owner", "Skill Name", "Full Name", "URL", "Description"]
CAT_COLS   = ["#", "Owner", "Skill Name", "Full Name", "URL", "Description"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def safe_sheet_name(name: str, used: set) -> str:
    """Truncate to Excel's 31-char limit, avoid duplicates."""
    cleaned = re.sub(r'[\\/*?:\[\]]', '-', name)[:31]
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f" ({n})"
        candidate = cleaned[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def style_header(ws, columns: list):
    ws.append(columns)
    for cell in ws[1]:
        cell.font = HEADER_FILL and HEADER_FONT
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 18


def autosize(ws, extra: dict = None):
    """Set column widths based on content, with optional overrides."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 3, 60)
        if extra and col_letter in extra:
            width = extra[col_letter]
        ws.column_dimensions[col_letter].width = width


def alternate_rows(ws, start_row: int = 2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        if i % 2 == 1:
            for cell in row:
                if cell.fill.patternType != "solid" or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = ALT_FILL


def add_hyperlink(cell, url: str):
    cell.value = url
    cell.hyperlink = url
    cell.font = URL_FONT


# ── Parser ────────────────────────────────────────────────────────────────────
def parse_skills(path: str) -> list[dict]:
    skills = []
    current_category = None
    in_microsoft = False
    in_community = False

    with open(path, encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.rstrip("\n")

        # ── Detect section headings ──────────────────────────────────────────
        h3_match = DETAILS_H3_RE.search(line) or PLAIN_H3_RE.match(line)
        if h3_match:
            raw = h3_match.group(1).strip()

            if raw == "Community Skills":
                in_community = True
                in_microsoft = False
                current_category = None
                continue

            if raw == "Skills by Microsoft":
                in_community = False
                in_microsoft = True
                current_category = raw
                continue

            if in_microsoft and raw in MICROSOFT_SUBSECTIONS:
                current_category = f"Microsoft – {raw}"
                continue

            if in_community:
                # Sub-category inside Community Skills
                current_category = f"Community – {raw}"
                continue

            # Regular top-level section
            in_microsoft = False
            in_community = False
            current_category = raw
            continue

        # ── Detect skill entries ─────────────────────────────────────────────
        skill_match = SKILL_RE.match(line)
        if skill_match and current_category:
            full_name = skill_match.group(1).strip()
            url       = skill_match.group(2).strip()
            desc      = skill_match.group(3).strip()

            # Strip any residual markdown/HTML from description
            desc = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', desc)
            desc = re.sub(r'<[^>]+>', '', desc)
            desc = desc.strip().rstrip('.')

            owner, sep, skill_name = full_name.partition("/")
            if not sep:
                skill_name = full_name
                owner = ""

            skills.append({
                "category":   current_category,
                "owner":      owner,
                "skill_name": skill_name,
                "full_name":  full_name,
                "url":        url,
                "description": desc,
            })

    return skills


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(skills: list[dict], output: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    style_header(ws_sum, ["Category", "Skill Count"])
    ws_sum.column_dimensions["A"].width = 50
    ws_sum.column_dimensions["B"].width = 14

    # Group by category preserving insertion order
    from collections import Counter, OrderedDict
    cat_counts: dict = OrderedDict()
    for s in skills:
        cat_counts.setdefault(s["category"], 0)
        cat_counts[s["category"]] += 1

    for i, (cat, count) in enumerate(cat_counts.items(), start=2):
        ws_sum.append([cat, count])
        if i % 2 == 1:
            for cell in ws_sum[i]:
                cell.fill = ALT_FILL
        ws_sum[f"B{i}"].alignment = Alignment(horizontal="center")

    # Total row
    total_row = ws_sum.max_row + 1
    ws_sum.append(["TOTAL", len(skills)])
    for cell in ws_sum[total_row]:
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    ws_sum[f"B{total_row}"].alignment = Alignment(horizontal="center")

    ws_sum.freeze_panes = "A2"

    # ── Sheet 2: All Skills ───────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Skills")
    style_header(ws_all, SHEET_COLS)

    for i, s in enumerate(skills, start=1):
        row_num = i + 1
        ws_all.append([
            i,
            s["category"],
            s["owner"],
            s["skill_name"],
            s["full_name"],
            s["url"],          # will be converted to hyperlink below
            s["description"],
        ])
        # Hyperlink on URL cell (column F = 6)
        url_cell = ws_all.cell(row=row_num, column=6)
        add_hyperlink(url_cell, s["url"])

        if i % 2 == 0:
            for col in range(1, 8):
                cell = ws_all.cell(row=row_num, column=col)
                if not (col == 6):  # don't override hyperlink font fill detection
                    cell.fill = ALT_FILL

    ws_all.freeze_panes = "A2"
    autosize(ws_all, extra={"F": 55, "G": 70, "B": 38})

    # ── Sheets 3–N: Per-category ──────────────────────────────────────────────
    used_names = {"Summary", "All Skills"}
    category_sheets: dict = {}

    for cat in cat_counts:
        sname = safe_sheet_name(cat, used_names)
        ws = wb.create_sheet(sname)
        style_header(ws, CAT_COLS)
        category_sheets[cat] = ws

    for i, s in enumerate(skills):
        ws = category_sheets[s["category"]]
        # Row index within this sheet
        row_num = ws.max_row + 1
        local_i = row_num - 1  # 1-based skill index within category

        ws.append([
            local_i,
            s["owner"],
            s["skill_name"],
            s["full_name"],
            s["url"],
            s["description"],
        ])
        url_cell = ws.cell(row=row_num, column=5)
        add_hyperlink(url_cell, s["url"])

        if local_i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = ALT_FILL

    for ws in category_sheets.values():
        ws.freeze_panes = "A2"
        autosize(ws, extra={"E": 55, "F": 70})

    wb.save(output)
    return cat_counts, len(skills)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"Parsing {README} …")
    skills = parse_skills(README)
    print(f"Found {len(skills)} skills.")

    print(f"Building {OUTPUT} …")
    cat_counts, total = build_excel(skills, OUTPUT)

    print(f"\n{'Category':<55} {'Count':>5}")
    print("-" * 62)
    for cat, cnt in cat_counts.items():
        print(f"  {cat:<53} {cnt:>5}")
    print("-" * 62)
    print(f"  {'TOTAL':<53} {total:>5}")
    print(f"\nSaved to {OUTPUT}")


if __name__ == "__main__":
    main()
